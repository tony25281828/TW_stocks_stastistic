'''
此程式可以搜尋想要的股票資訊，並且將資料匯入csv file裏面
可搜索的資訊包含
1. 某日的股票訊息
2. 某月的股票訊息，並且繪製圖表
3. 某年的股票訊息，並且繪製圖表
4. 一段時間的股票訊息，並且繪製圖表

5. 儲存到excel
6. 儲存到mysql
7. 傳送email
8. 新增了bookmark外掛程式測試
'''

import datetime
import os
import matplotlib.pyplot as plt
import requests
import json
import pandas as pd
import csv
import sys
from plotly.offline import plot
from plotly.graph_objs import Scatter, Layout
from tabulate import tabulate
import numpy
import pyinputplus as pyip
from datetime import datetime
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.chart import LineChart, Reference
from pathlib import Path



# 因為我下載之後利用panda去修改裏面的資料，可能會有chain的問題
# 所以跳出warning，所以用這一行代碼來ignore warning
pd.options.mode.chained_assignment = None



# 獲取整年的urls
# YYYY(string)只有年份
def get_year_url(stockNumber, year):
    fullUrlList = []
    headUrl = 'https://www.twse.com.tw/exchangeReport/STOCK_DAY?response=json&date='
    if int(year) < datetime.now().year:
        for monthNum in range(12):
            if monthNum < 9:
                fullUrl = headUrl + year + '0' + str(monthNum+1) + '01&stockNo=' + stockNumber
            else:
                fullUrl = headUrl + year + str(monthNum + 1) + '01&stockNo=' + stockNumber
            fullUrlList.append(fullUrl)
    else:
        for monthNum in range(datetime.now().month):
            if monthNum < 9:
                fullUrl = headUrl + year + '0' + str(monthNum+1) + '01&stockNo=' + stockNumber
            else:
                fullUrl = headUrl + year + str(monthNum + 1) + '01&stockNo=' + stockNumber
            fullUrlList.append(fullUrl)
    return fullUrlList





# 獲取某一年到某一年的urls
# start_YYYY以及end_YYYY只有年份(string)
def get_years_url(stockNumber, start_year, end_year):
    fullUrlList = []
    yearList = list(range(int(start_year), int(end_year)+1))
    for year in yearList:
        fullUrlList += get_year_url(stockNumber, str(year))
    return fullUrlList





# 檢查我們要尋找的年份資料是否已經存在檔案裡了
# startDate跟endDate是包含年月日的string
# return 0, 1, 2, 3, 4
# 0 但表完全沒檔案
# 1 代表資料在資料檔案裡，或是已是最新檔案
# 2 代表有今年的資料，但是尚未更新到最新
# 3 代表有資料檔案，但是沒有想要查詢的資料
# 4 代表缺少資料
# 5 代表無法進行網路連結
# 分成三種判斷
    # 第一種：如果開始年份小於檔案名稱裡的開始年份或是結束年份大於檔案名稱裏面的結束年份
    # 第二種：結束年份在今年以前，只需要查詢歷史資料即可
    # 第三種：結束年份為今年，先檢查是否有最新資料，有最新資料再去檢查有沒有我們想要的資料。如果沒有最新資料，必須先更新
def is_info_exists(csv_file_path, startDate, endDate):
    # 0 代表沒有檔案，必須建立新檔案
    exist = 0
    if csv_file_path != '':
        try:
            data = pd.read_csv(csv_file_path)
        except:
            print('找不到檔案路徑')
            return exist
        # 初始化年份資料
        csv_file_name = os.path.basename(csv_file_path)
        csv_file_name_startyear = int(csv_file_name[5:9])
        csv_file_name_endyear = int(csv_file_name[10:14])
        start_year = int(startDate[:4])
        end_year = int(endDate[:4])
    # 如果開始或結束年都在file的開始年份前面或是結束年份後面，代表有資料檔案，但是要查詢的資料不在裡面
        if start_year < csv_file_name_startyear or end_year > csv_file_name_endyear:
            print(csv_file_name + ' 資料檔案中沒有想要查詢的資料')
            exist = 3
    # 如果結束年份是在今年以前，只需要檢查歷史資料有沒有完整，以及檢查要查詢的年份有沒有在檔案資料裏面
        elif end_year < datetime.now().year:
            # 將資料檔案有的年份資料加入set裏面
            years_in_data = set()
            for year_in_data in data['日期']:
                years_in_data.add(int(year_in_data[:4]))
            # 要查詢的年份
            search_list = range(start_year, end_year+1)
            lost_years = []
            # 開始查詢
            for year in search_list:
                if year in years_in_data:
                    # 如果要查詢的年份有在檔案資料裡，就返回1
                    exist = 1
                    print('想查詢的資料已在' + csv_file_name)
                else:
                    # 將缺失的年份加入list
                    lost_years.append(year)
            if len(lost_years) != 0:
                for lost_year in lost_years:
                    # 如果中間有一年的年份沒有載檔案資料裡，就返回3，並且break
                    print('缺失了' + str(lost_year) + '年的資料')
                exist = 4

    # 如果要查詢的結束年份為今年，先查詢今年的資料是否為最新資料
    # 如果是最新資料，再去檢查資料檔案裡是否有我們要的資料
    # 如果不是最新資料，要先把今年的資料更新到最新
        elif end_year == datetime.now().year:
            # 獲取資料檔案裡面的最後一筆資料日期
            the_latest_date_csv_file = data['日期'].tail(1).values[0]
            # 如果年份包含今年，上台灣證券網站上面或去最新月份的最後一日日期
            head_url = 'https://www.twse.com.tw/exchangeReport/STOCK_DAY?response=json&date='
            tail_url = '&stockNo='
            stock_num = os.path.basename(csv_file_path)[:4]
            today_date = datetime.strftime(datetime.now(), '%Y%m%d')
            stockUrl = head_url + today_date + tail_url + stock_num
            try:
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36'
                }
                req = requests.get(stockUrl, headers=headers)
            except:
                print('無法開啟網路連結，請稍後再試')
                exist = 5
                return exist
            root = json.loads(req.text)
            # 獲取網站上最後一天的日期，並且轉換成int
            the_latest_date_web = root['data'][-1][0]
            y = str(int(the_latest_date_web[:3]) + 1911)
            the_latest_date_web = y + '-' + the_latest_date_web[4:6] + '-' + the_latest_date_web[7:]
            # 資料檔案裡面最後一天的資料等於網站上最後一天資料代表今年的資料是最新
            # 並且查詢想要查詢的年份是否在檔案資料裏面
            if the_latest_date_csv_file == the_latest_date_web:
                years_in_data = set()
                for year_in_data in data['日期']:
                    years_in_data.add(int(year_in_data[:4]))
                search_list = range(start_year, end_year + 1)
                for year in search_list:
                    if year in years_in_data:
                        # 如果要查詢的年份有在檔案資料裡，就返回1
                        exist = 1
                    else:
                        # 如果中間有一年的年份沒有載檔案資料裡，就返回3，並且break
                        exist = 3
                        break
                if exist == 1:
                    print(csv_file_name + '有' + str(end_year) + '年最新資料以及要查詢的年份資料')
                    print('=' * 30)
            else:
                # 資料檔案裏面有今年的資料，但是今年資料尚未更新至最新 exist = 2
                exist = 2
                print('請將' + str(end_year) + '年的資料更新到最新')
    # 如果檔案路徑是空的，代表完全沒有檔案，因此返回0準備建檔
    else:
        print('資料檔案庫裏面沒有資料，準備建立檔案及下載資料')
        exist = 0
    return exist





# 下載股票資料，並且轉換成pandas.DataFrame Obj
def convert_stocksInfo_to_pd(stockUrls):
    # 新增這個list是為了要將所有擷取的資料加入這個list，之後才好加入pandas.DataFrame裏面
    stockInfo_list = []
    for stockUrl in stockUrls:
        year = stockUrl[stockUrl.index('date=') + 5:stockUrl.index('date=') + 9]
        month = stockUrl[stockUrl.index('date=') + 9:stockUrl.index('date=') + 11]
        stock_num = stockUrl[-4:]
       # request URL
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36'
            }
            req = requests.get(stockUrl, headers=headers)
        except:
            print('網路無法連結，請稍後再試')
            return pd.DataFrame()
        # 解讀json資料
        root = json.loads(req.text)
        # 從解讀後的json資料或去想要的資訊
        try:
            data = root['data']
        except:
            print('成功連上網路，但是無資料擷取')
            return pd.DataFrame()
        print('正在下載' + str(stock_num) + ' ' + str(year) + '年' + str(month) + '月的資料')
        for x in data:
            stockInfo_list.append(x)
    stocksInfo_df = pd.DataFrame(stockInfo_list)
    return stocksInfo_df





# 將下載好的pandas.DataFrame更新到csv檔案裡面
def update_csv(csv_file_path, pandaDataFrame):
    # 寫入csv，使用'a'模式，可以從最後面新加資料上去，不會被覆寫過去
    csvFile = open(csv_file_path, 'a', newline='')
    csvFileWriter = csv.writer(csvFile)
    for dataNum in range(pandaDataFrame.index.stop):
        data_daily = pandaDataFrame.iloc[dataNum,:].tolist()
        csvFileWriter.writerow(data_daily)
    csvFile.close()
    return print('資料成功加入資料檔案')




# 將某個csv檔案加入想要的header
def csv_add_header(csvFilePath, header):
    data = pd.read_csv(csvFilePath)
    if data.columns.all() != header:
        data.to_csv(csvFilePath, header=header, index=False)






# 將某個csv檔案裏面的民國日期轉換為西元的格式
def convertDate(csvFilePath):
    data = pd.read_csv(csvFilePath)
    for x in range(len(data['日期'])):
        if str(data['日期'][x]).__contains__('/'):
            year = str(int(data['日期'][x][0:3]) + 1911)
            month = data['日期'][x][4:6]
            date = data['日期'][x][7:]
            newDateStr = year + '-' + month + '-' + date
            data['日期'][x] = newDateStr
    data.to_csv(csvFilePath, index=False)





# 將csv檔案裏面的資料通通轉換成int或是float格式
def convert_to_int_in_csv(csvFile):
    data = pd.read_csv(csvFile)
    for x in range(len(data)):
        #print(type(data['成交股數'][x]))
        # ininstance(Obj, type) 可以檢查是否是某個type
        if not isinstance(data['成交股數'][x], numpy.int64):
            data['成交股數'][x] = int(data['成交股數'][x].replace(',',''))
        if not isinstance(data['成交金額'][x], numpy.int64):
            data['成交金額'][x] = int(data['成交金額'][x].replace(',', ''))
        if not isinstance(data['成交筆數'][x], numpy.int64):
            data['成交筆數'][x] = int(data['成交筆數'][x].replace(',', ''))
        if not isinstance(data['漲跌價差'][x], numpy.float64):
            if data['漲跌價差'][x].__contains__('+'):
                data['漲跌價差'][x] = float(data['漲跌價差'][x].replace('+',''))
            elif data['漲跌價差'][x].__contains__('-'):
                data['漲跌價差'][x] = 0 - float(data['漲跌價差'][x].replace('-',''))
    data.to_csv(csvFile, index=False)





# 如果檔案裡面沒有我們要的資料，開始擷取台灣證券交易所的資料
# 根據isintoexists所回傳的數值來做適當的處理
# start_year跟end_year是只有year(string)
def download_stocks_data(isinfoexists, stock_file_dir, stock_num, start_year, end_year, csv_file_path):
    csv_file_name = os.path.basename(csv_file_path)
    start_year_name = csv_file_name[5:9]
    end_year_name = csv_file_name[10:14]
    # 完全沒有檔案，建立新的檔案
    if isinfoexists == 0:
        start_year_name = str(start_year)
        end_year_name = str(end_year)
        print('準備下載')
        csv_file_path = stock_file_dir + '/' + stock_num + '_' + start_year_name + '_' + end_year_name + '.csv'
        # 獲取每月的股票url
        stocks_url_list = get_years_url(stock_num, start_year, end_year)
        # 將每一筆資料轉化成pandas.DataFrame
        df = convert_stocksInfo_to_pd(stocks_url_list)
        if df.size == 0:
            return print('資料加入資料檔案失敗，請稍後再試')
        # 將pandas.DataFrame資料更新到csv檔案
        update_csv(csv_file_path, df)
        # 加入header
        header = ['日期', '成交股數', '成交金額', '開盤價', '最高價', '最低價', '收盤價', '漲跌價差', '成交筆數']
        csv_add_header(csv_file_path, header)

    # 有最新檔案或是檔案已在資料檔案裡，不需做任何動作
    elif isinfoexists == 1:
        return csv_file_path

    # 有今年的檔案，但是尚未更新至最新，因此更新檔案內容即可，不需更新檔案名稱
    elif isinfoexists == 2:
        csv_file_name = os.path.basename(csv_file_path)

        # 獲取每月的股票url
        stocks_url_list = get_years_url(stock_num, end_year, end_year)
        # 整理一下stocks_url_list，刪除不必要的月份
        del stocks_url_list[datetime.now().month:]
        print('開始更新今年的資料')
        # 將每一筆資料轉化成pandas.DataFrame
        df = convert_stocksInfo_to_pd(stocks_url_list)
        if df.size == 0:
            return print('資料加入資料檔案失敗，請稍後再試')
        # 將pandas.DataFrame資料更新到csv檔案
        update_csv(csv_file_path, df)

    # 檔案資料裡面有我們要的股票代號，但是沒有我們要的資料年份，所以更新年份，同時更新檔案名稱
    elif isinfoexists == 3:
        csv_file_name = os.path.basename(csv_file_path)

        # 根據要開始年份、結束年份以及現有的檔案名稱來重新命名
        if int(start_year) < int(csv_file_name[5:9]):
            start_year_name = start_year
        if int(end_year) > int(csv_file_name[10:14]):
            end_year_name = end_year

        # 根據要搜尋的開始年份跟結束年份還有現有的檔案年份來定義要下載的年份
        if int(csv_file_name[10:14]) >= int(end_year) >= int(csv_file_name[5:9]):
            end_year = int(csv_file_name[5:9]) - 1
        if int(csv_file_name[10:14]) >= int(start_year) >= int(csv_file_name[5:9]):
            start_year = int(csv_file_name[10:14]) + 1

        # 獲取每月的股票url
        stocks_url_list = get_years_url(stock_num, start_year, end_year)
        print('開始下載')
        # 將每一筆資料轉化成pandas.DataFrame
        df = convert_stocksInfo_to_pd(stocks_url_list)
        if df.size == 0:
            return print('資料加入資料檔案失敗，請稍後再試')
        # 更新檔案名稱
        new_csv_file_path = stock_file_dir + '/' + csv_file_name[:4] + '_' + str(start_year_name) + '_' + str(end_year_name) + '.csv'
        os.rename(csv_file_path, new_csv_file_path)
        csv_file_path = new_csv_file_path
        # 將pandas.DataFrame資料更新到csv檔案
        update_csv(csv_file_path, df)

    # 有缺失的資料，只需下載遺失資料即可，不需更新檔案名稱
    elif isinfoexists == 4:
        # 獲取每月的股票url
        stocks_url_list = get_years_url(stock_num, start_year, end_year)
        print('開始下載')
        # 將每一筆資料轉化成pandas.DataFrame
        df = convert_stocksInfo_to_pd(stocks_url_list)
        if df.size == 0:
            return print('資料加入資料檔案失敗')
        # 將pandas.DataFrame資料更新到csv檔案
        update_csv(csv_file_path, df)

    # 將日期轉換成西元格式
    convertDate(csv_file_path)

    # 將擷取過後的資料通通轉換成int
    convert_to_int_in_csv(csv_file_path)
    # 將資料依照順序排列
    data_unsorted = pd.read_csv(csv_file_path)
    date_sorted_by_date = data_unsorted.sort_values(by=['日期'])
    date_sorted_by_date.drop_duplicates(keep='first', inplace=True)
    date_sorted_by_date.to_csv(csv_file_path, index=False)

    print('資料已儲存在 ' + os.path.basename(csv_file_path))
    print('=' * 30)
    return csv_file_path





# 查詢某一天的資料，startDate跟endDate=''是包含年月日的string
# 查詢某一段時間的資料，startDate跟endDate是包含年月日的string
def get_data_by_date(csv_file_path, startDate, endDate=''):
    if endDate == '':
        try:
            data = pd.read_csv(csv_file_path, index_col='日期')
            startDate = startDate[:4] + '-' + startDate[4:6] + '-' + startDate[6:]
            print('日期: ', startDate)
            try:
                print(data.loc[startDate, :])
            except:
                print('資料檔案中無資料，請回到主選單選擇2，新增年份資料。\n'
                      '如果資料檔案裡已有年份資料，代表當天為國定假日或是週末，並未開盤。')
                return
        except:
            print('查無資料')
            return
    else:
        data = pd.read_csv(csv_file_path)
        endDate = endDate[:4] + '-' + endDate[4:6] + '-' + endDate[6:]
        data['日期'] = pd.to_datetime(data['日期'], format='%Y-%m-%d')
        df = data[(data['日期'] <= endDate) & (data['日期'] >= startDate)]
        #print(tabulate(df, headers='keys', tablefmt='grid'))
        return df





# 畫出開盤價、最高價、最低價、收盤價靜態折線圖
def plot_line_chart(pandaDataFrame, stock_num):
    # 開始畫圖
    # 設定中文字型
    plt.rcParams['font.sans-serif'] = 'Arial Unicode MS'
    # 設定負號正確顯示
    plt.rcParams['axes.unicode_minus'] = False

    # 畫出開盤價、最高價、最低價、收盤價
    pandaDataFrame.plot(kind='line', figsize=[20,10], x='日期', y=['開盤價', '最高價','最低價','收盤價'],grid=True)
    plt.title(os.path.basename(stock_num)[:4] + '股價走勢圖', size=30)
    plt.xlabel('日期 (YYYY-MM)', size=20, color='blue')
    plt.ylabel('價格', size=20, color='blue')

    '''
    pandaDataFrame.plot(kind='line', figsize=[20,10], x='日期', y='成交金額',grid=True, color='green')
    plt.title(stock_num + '成交金額圖', size=20)
    plt.xlabel('日期 (YYYY-MM)', size=20, color='blue')
    plt.ylabel('金額', size=20, color='blue')

    pandaDataFrame.plot(kind='line', figsize=[20,10], x='日期', y='成交筆數',grid=True, color='red')
    plt.title(stock_num + '成交筆數圖', size=20)
    plt.xlabel('日期 (YYYY-MM)', size=20, color='blue')
    plt.ylabel('筆數', size=20, color='blue')

    pandaDataFrame.plot(kind='line', figsize=[20, 10], x='日期', y='成交股數', grid=True, color='blue')
    plt.title(stock_num + '成交股數圖', size=20)
    plt.xlabel('日期 (YYYY-MM)', size=20, color='blue')
    plt.ylabel('股數', size=20, color='blue')
    '''

    plt.show()
    print(tabulate(pandaDataFrame, headers='keys', tablefmt='grid'))





# 畫出開盤價、最高價、最低價、收盤價動態html折線圖
def plot_dynamic_chart(pandaDataFrame, stock_num):
    # 開始畫圖
    # 設定中文字型
    plt.rcParams['font.sans-serif'] = 'Arial Unicode MS'
    # 設定負號正確顯示
    plt.rcParams['axes.unicode_minus'] = False

    plot_data = [
        Scatter(x=pandaDataFrame['日期'], y=pandaDataFrame['開盤價'], name='開盤價'),
        Scatter(x=pandaDataFrame['日期'], y=pandaDataFrame['最高價'], name='最高價'),
        Scatter(x=pandaDataFrame['日期'], y=pandaDataFrame['最低價'], name='最低價'),
        Scatter(x=pandaDataFrame['日期'], y=pandaDataFrame['收盤價'], name='收盤價')
    ]
    plot({'data':plot_data, 'layout':Layout(title=os.path.basename(stock_num)[:4] +'股價統計圖')},
         auto_open=True)
    print(tabulate(pandaDataFrame, headers='keys', tablefmt='grid'))





# 將資料儲存為excel檔案
def save_to_excel(pandaDataFrame, stock_num, start_date, end_date, path):
    # 建立一個workbook物件
    excelObj = openpyxl.Workbook()
    # 設定要編輯的sheet，並且設定一些參數
    sheet = excelObj.active
    sheet.title = stock_num
    sheet['A1'] = stock_num
    header = list(pandaDataFrame.columns)
    # 加入header
    sheet.append(header)
    # pandaDataFrame.values是一個numpy array，所以用tolist()轉成list
    # 加入value
    pandaDataFrame['日期'] = pd.to_datetime(pandaDataFrame['日期'],format='%Y%m%d')
    values = pandaDataFrame.values.tolist()
    for value in values:
        sheet.append(value)

    # 設定cell的寬度
    for x in range(len(header)):
        if x <= 2:
            sheet.column_dimensions[get_column_letter(x+1)].width = 22
        else:
            sheet.column_dimensions[get_column_letter(x+1)].width = 12
    # 設定字體大小
    font_size = Font(size=13)
    for x in range(sheet.max_column):
        for y in range(sheet.max_row):
            sheet.cell(y+1,x+1).font = font_size

    # 開始畫圖
    lineChart = LineChart()
    lineChart.title = stock_num + '股票折線圖'
    lineChart.width = 40
    lineChart.height = 25
    # 設定x軸的最大值跟最小值
    lineChart.y_axis.scaling.min = float(pandaDataFrame['最低價'].values.min())*0.9
    lineChart.y_axis.scaling.max = float(pandaDataFrame['最高價'].values.max())*1.1
    # 將資料加入reference物件
    data1 = Reference(sheet, min_col=4, min_row=3, max_row=sheet.max_row)
    data2 = Reference(sheet, min_col=5, min_row=3, max_row=sheet.max_row)
    data3 = Reference(sheet, min_col=6, min_row=3, max_row=sheet.max_row)
    data4 = Reference(sheet, min_col=7, min_row=3, max_row=sheet.max_row)
    dates = Reference(sheet, min_col=1, min_row=3, max_row=sheet.max_row)
    # 將reference物件建立成series物件
    series_data1 = openpyxl.chart.Series(data1, title='開盤價')
    series_data2 = openpyxl.chart.Series(data2, title='最高價')
    series_data3 = openpyxl.chart.Series(data3, title='最低價')
    series_data4 = openpyxl.chart.Series(data4, title='收盤價')
    # 設定x軸跟y軸
    lineChart.x_axis.number_format = 'YYYYMMDD'
    lineChart.x_axis.majorTimeUnit = 'days'
    lineChart.x_axis.title = '日期'
    lineChart.y_axis.title = '股價'
    #將series物件加入圖裏面
    lineChart.append(series_data1)
    lineChart.append(series_data2)
    lineChart.append(series_data3)
    lineChart.append(series_data4)
    lineChart.set_categories(dates)
    sheet.add_chart(lineChart, 'M5')

    excelObj.save(path + str(stock_num) + '_' + str(start_date) + '_' + str(end_date) + '.xlsx')
    print('已將excel檔案儲存到桌面')
    return





# 是否要來開程式
def exist_system():
    input = pyip.inputYesNo('要繼續服務(y)或是離開(n)?\n')
    if input == 'no':
        print('掰掰～～～')
        sys.exit()





# 更新stocks_data的資料，沒有放在程式裏面，需要自己手動更新
def update_stocks_data():
    dir = './stocks_data_file'
    csv_file_name = ''
    for folder, subfolder, fileNames in os.walk(dir):
        for fileName in fileNames:
            if fileName.__contains__('stocks_data') and fileName.endswith('.csv'):
                csv_file_name = fileName
    

    url = 'https://isin.twse.com.tw/isin/C_public.jsp?strMode=2'

    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/92.0.4515.131 Safari/537.36'
        }
        req = requests.get(url, headers=headers)
    except:
        print('無法連上伺服器')
    root = BeautifulSoup(req.text, 'lxml')
    time_title = root.findAll('h2')[1].text
    latest_date = time_title[7:].replace('/','').rstrip()


    if csv_file_name == '':
        print('資料檔案無資料，正在下載...')
        csv_file_name = dir + '/' + 'stocks_data_' + latest_date + '.csv'
        csv_file = open(csv_file_name, 'w', newline='')
        header = ['股票代號','公司名稱']
        csv_file_Dictwriter = csv.DictWriter(csv_file, fieldnames=header)
        csv_file_Dictwriter.writeheader()
        titles = root.findAll('tr')
        for titleNum in range(2, len(titles)):
            # chr(12288)是中文的空格，' '是找不到的
            title = titles[titleNum].td.text.replace(chr(12288),',')
            try:
                stocks_num = '/' + title[:title.index(',')]
                stocks_name = title[title.index(',')+1:]
                csv_file_Dictwriter.writerow({'股票代號':stocks_num, '公司名稱':stocks_name})
            except:
                pass
    else:
        file_date_str = csv_file_name[12:20]
        file_date_datetime = datetime.strptime(file_date_str, '%Y%m%d')
        latest_date_datetime = datetime.strptime(latest_date, '%Y%m%d')
        if file_date_datetime == latest_date_datetime:
            print('股票公司名稱資料已是最新，不需更新')
        else:
            print('資料檔案更新中...')
            csv_file = open(dir + '/' + csv_file_name, 'w', newline='')
            header = ['股票代號', '公司名稱']
            csv_file_Dictwriter = csv.DictWriter(csv_file, fieldnames=header)
            csv_file_Dictwriter.writeheader()
            titles = root.findAll('tr')
            for titleNum in range(2, len(titles)):
                # chr(12288)是中文的空格，' '是找不到的
                title = titles[titleNum].td.text.replace(chr(12288), ',')
                try:
                    stocks_num = '/' + title[:title.index(',')]
                    stocks_name = title[title.index(',') + 1:]
                    csv_file_Dictwriter.writerow({'股票代號': stocks_num, '公司名稱': stocks_name})
                except:
                    pass
            new_csv_file_name = dir + '/' + csv_file_name[:12] + latest_date + '.csv'
            csv_file_name = dir + '/' + csv_file_name
            print(new_csv_file_name)
            os.rename(csv_file_name, new_csv_file_name)
            csv_file.close()





# 利用股票代號查詢股票，或是公司名稱查詢股票代號
def search_stocks(stock_num_or_name):
    # 先找出資料檔案
    dir = './stocks_data_file/'
    csv_file_name = ''
    for folder, subfolder, fileNames in os.walk(dir):
        for fileName in fileNames:
            if fileName.__contains__('stocks_data') and len(fileName) == 24:
                csv_file_name = fileName
    # 用panda讀取檔案
    csv_file = pd.read_csv(dir + csv_file_name)
    # 用dataFrame開始尋找股票代號或是公司名稱
    results = []
    for x in range(len(csv_file)):
        if str(csv_file.iloc[x,0].replace('/','')) == stock_num_or_name:
            results.append(csv_file.iloc[x,0].replace('/','') + " " + csv_file.iloc[x,1])
            break
        elif csv_file.iloc[x,1].__contains__(stock_num_or_name):
            results.append(csv_file.iloc[x, 0].replace('/', '') + " " + csv_file.iloc[x, 1])

    for result in results:
        print(result)
    print('總共: ' + str(len(results)) + '個結果')


# 更新股票代號跟公司名稱資料庫，需要手動更新
#update_stocks_data()